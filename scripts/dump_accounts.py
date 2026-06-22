import openpyxl
path = r"C:\Users\User\Downloads\任務追蹤系統_測試仕樣書_v1.1_待測_20260602.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['② 測試帳號與資料']
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(row):
        print(r, row)
