"""
Write back test results to a new copy of the test-spec xlsx.
- Reads original from C:\\Users\\User\\Downloads\\任務追蹤系統_測試仕樣書_v1.1_待測_20260602.xlsx
- Writes to    C:\\Users\\User\\Downloads\\任務追蹤系統_測試仕樣書_v1.1_測完_20260604.xlsx
- ④測試情境 sheet, column N=實際結果(API+UI合併), O=判定
"""
import json, shutil, sys
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\User\Downloads\任務追蹤系統_測試仕樣書_v1.2_待測_20260606.xlsx")
DST = Path(r"C:\Users\User\Downloads\任務追蹤系統_測試仕樣書_v1.2_測完_20260608.xlsx")
API_RESULTS = Path(__file__).parent / "test_results.json"
UI_RESULTS  = Path(__file__).parent / "ui_results.json"
SCENARIOS   = Path(__file__).parent / "scenarios.json"

if not SRC.exists():
    print(f"[ERR] source xlsx not found: {SRC}"); sys.exit(1)

shutil.copy(SRC, DST)

api_results = {r["id"]: r for r in json.loads(API_RESULTS.read_text(encoding="utf-8"))}
ui_results  = {r["id"]: r for r in json.loads(UI_RESULTS.read_text(encoding="utf-8"))}
scenarios   = json.loads(SCENARIOS.read_text(encoding="utf-8"))
tc_to_row   = {s["id"]: s["row"] for s in scenarios if str(s.get("id", "")).startswith("TC-")}

wb = load_workbook(DST)
sheet_name = next((n for n in wb.sheetnames if "情境" in n), None)
if not sheet_name:
    print("[ERR] no scenario sheet"); sys.exit(1)
ws = wb[sheet_name]

JUDGE = {"PASS": "通過", "FAIL": "不通過", "SKIP": "未測"}

def merge_status(a, u):
    """FAIL > PASS > SKIP"""
    if a == "FAIL" or u == "FAIL": return "FAIL"
    if a == "PASS" or u == "PASS": return "PASS"
    return "SKIP"

count = 0
for tc_id, row in tc_to_row.items():
    api_r = api_results.get(tc_id)
    ui_r  = ui_results.get(tc_id)
    if not api_r and not ui_r: continue

    api_status = api_r["status"] if api_r else "SKIP"
    ui_status  = ui_r["status"]  if ui_r  else "SKIP"
    api_note   = (api_r["note"][:200] if api_r.get("note") else "") if api_r else ""
    ui_note    = (ui_r["note"][:200]  if ui_r.get("note")  else "") if ui_r  else ""

    combined_status = merge_status(api_status, ui_status)
    parts = []
    if api_note: parts.append(f"API:{api_note}")
    if ui_note:  parts.append(f"UI:{ui_note}")
    actual_text = f"[API:{api_status}|UI:{ui_status}] " + " / ".join(parts)

    ws.cell(row=row, column=14).value = actual_text[:500]  # N 實際結果
    ws.cell(row=row, column=15).value = JUDGE.get(combined_status, combined_status)  # O 判定
    count += 1

wb.save(DST)
print(f"[OK] wrote {count} rows to {DST}")
