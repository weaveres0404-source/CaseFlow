import openpyxl, json
path = r"C:\Users\User\Downloads\任務追蹤系統_測試仕樣書_v1.1_待測_20260602.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['④ 測試情境']
out = []
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if row[10] or row[11] or row[12]:
        out.append({
            'row': r, 'id': row[0], 'category': row[1], 'name': row[2],
            'priority': row[3], 'precondition': row[4], 'role': row[5],
            'steps': row[6], 'input': row[7], 'expectedUI': row[8],
            'stateTrans': row[9], 'method': row[10], 'endpoint': row[11],
            'expectedHttp': row[12], 'actualResult': row[13], 'judgement': row[14],
        })
with open(r'D:\CaseFlow\CaseFlow\scripts\scenarios.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"Total scenarios with API spec: {len(out)}")
