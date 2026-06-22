import openpyxl, json, sys, os
path = r"C:\Users\User\Downloads\任務追蹤系統_測試仕樣書_v1.1_待測_20260602.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("SHEETS:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== {name} ({ws.max_row} rows x {ws.max_column} cols) ===")
    for r in range(1, min(ws.max_row, 4) + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        print(r, row)
